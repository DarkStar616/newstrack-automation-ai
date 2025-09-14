"""
Excel ingestion utility for loading keywords with source location data.
Supports .xls/.xlsx files with fallback across different Excel libraries.
"""
import os
import io
from typing import List, Dict, Optional, Any, Union
import logging


def load_keywords(file_path: str) -> List[Dict[str, Optional[str]]]:
    """
    Load keywords from Excel file with source location data.
    
    Args:
        file_path: Path to .xls or .xlsx file
        
    Returns:
        List of dictionaries with keys:
        - keyword: The keyword string
        - category: Optional category (company/industry/regulatory)  
        - source_location: Optional source location rule
        
    Expected Excel columns:
    - Keyword (required)
    - Category (optional: company/industry/regulatory)
    - Source location (optional: blank | "South Africa" | "!South Africa")
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
        
    # Try different Excel reading approaches
    try:
        return _load_with_pandas(file_path)
    except ImportError:
        logging.warning("pandas not available, trying openpyxl")
        try:
            return _load_with_openpyxl(file_path)
        except ImportError:
            logging.warning("openpyxl not available, trying xlrd")
            try:
                return _load_with_xlrd(file_path)
            except ImportError:
                raise ImportError(
                    "No Excel reading library available. "
                    "Please install pandas, openpyxl, or xlrd"
                )


def _load_with_pandas(file_path: str) -> List[Dict[str, Optional[str]]]:
    """Load Excel using pandas (preferred method)."""
    import pandas as pd
    
    # Read the Excel file
    if file_path.endswith('.xlsx'):
        df = pd.read_excel(file_path, engine='openpyxl')
    else:
        df = pd.read_excel(file_path, engine='xlrd')
    
    return _process_dataframe(df)


def _load_with_openpyxl(file_path: str) -> List[Dict[str, Optional[str]]]:
    """Load Excel using openpyxl (xlsx only)."""
    if not file_path.endswith('.xlsx'):
        raise ValueError("openpyxl only supports .xlsx files")
        
    from openpyxl import load_workbook
    
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active
    
    # Get headers from first row
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Process data rows
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = value
        data.append(row_dict)
    
    # Convert to our format
    return _process_raw_data(data)


def _load_with_xlrd(file_path: str) -> List[Dict[str, Optional[str]]]:
    """Load Excel using xlrd (older .xls files)."""
    import xlrd
    
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)
    
    # Get headers from first row
    headers = []
    for col in range(sheet.ncols):
        headers.append(sheet.cell_value(0, col))
    
    # Process data rows
    data = []
    for row in range(1, sheet.nrows):
        row_dict = {}
        for col in range(sheet.ncols):
            if col < len(headers) and headers[col]:
                row_dict[headers[col]] = sheet.cell_value(row, col)
        data.append(row_dict)
    
    return _process_raw_data(data)


def _process_dataframe(df) -> List[Dict[str, Optional[str]]]:
    """Process pandas DataFrame into our format."""
    results = []
    
    # Normalize column names (case-insensitive matching)
    columns = {col.lower().strip(): col for col in df.columns}
    
    keyword_col = _find_column(columns, ['keyword', 'keywords', 'term', 'terms'])
    category_col = _find_column(columns, ['category', 'type', 'classification'])
    source_location_col = _find_column(columns, ['source location', 'source_location', 'region', 'location'])
    
    if not keyword_col:
        raise ValueError("No keyword column found. Expected columns: Keyword, Category, Source location")
    
    for _, row in df.iterrows():
        keyword = str(row[keyword_col]).strip() if not pd.isna(row[keyword_col]) else None
        
        if not keyword or keyword.lower() in ['nan', 'none', '']:
            continue
            
        category = None
        if category_col and not pd.isna(row[category_col]):
            category = str(row[category_col]).strip()
            if category.lower() in ['nan', 'none', '']:
                category = None
        
        source_location = None  
        if source_location_col and not pd.isna(row[source_location_col]):
            source_location = str(row[source_location_col]).strip()
            if source_location.lower() in ['nan', 'none', '']:
                source_location = None
        
        results.append({
            'keyword': keyword,
            'category': _normalize_category(category),
            'source_location': source_location
        })
    
    return results


def _process_raw_data(data: List[Dict]) -> List[Dict[str, Optional[str]]]:
    """Process raw data dictionaries into our format."""
    import pandas as pd
    
    # Convert to DataFrame for consistent processing
    df = pd.DataFrame(data)
    return _process_dataframe(df)


def _find_column(columns: Dict[str, str], candidates: List[str]) -> Optional[str]:
    """Find a column by trying multiple candidate names."""
    for candidate in candidates:
        if candidate in columns:
            return columns[candidate]
    return None


def _normalize_category(category: Optional[str]) -> Optional[str]:
    """Normalize category values to standard format."""
    if not category:
        return None
        
    category = category.lower().strip()
    
    # Map variations to standard categories
    if category in ['company', 'companies', 'corp', 'corporation']:
        return 'company'
    elif category in ['industry', 'industries', 'sector', 'business']:
        return 'industry' 
    elif category in ['regulatory', 'regulation', 'regulator', 'compliance']:
        return 'regulatory'
    else:
        return category


def validate_excel_format(file_path: str) -> Dict[str, Any]:
    """
    Validate Excel file format and return info about the file.
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        Dictionary with validation results:
        - valid: bool
        - row_count: int
        - columns_found: list
        - has_keyword_column: bool
        - has_category_column: bool
        - has_source_location_column: bool
        - sample_rows: list (first 3 rows)
    """
    try:
        keywords = load_keywords(file_path)
        
        # Try to reload with pandas for column analysis
        try:
            import pandas as pd
            if file_path.endswith('.xlsx'):
                df = pd.read_excel(file_path, engine='openpyxl')
            else:
                df = pd.read_excel(file_path, engine='xlrd')
                
            columns = [col.lower().strip() for col in df.columns]
            
            return {
                'valid': True,
                'row_count': len(keywords),
                'columns_found': list(df.columns),
                'has_keyword_column': any(col in columns for col in ['keyword', 'keywords', 'term', 'terms']),
                'has_category_column': any(col in columns for col in ['category', 'type', 'classification']),
                'has_source_location_column': any(col in columns for col in ['source location', 'source_location', 'region', 'location']),
                'sample_rows': keywords[:3]
            }
        except:
            return {
                'valid': True,
                'row_count': len(keywords),
                'columns_found': ['Unknown'],
                'has_keyword_column': True,  # Must be true if we loaded successfully
                'has_category_column': False,
                'has_source_location_column': False,
                'sample_rows': keywords[:3]
            }
            
    except Exception as e:
        return {
            'valid': False,
            'error': str(e),
            'row_count': 0,
            'columns_found': [],
            'has_keyword_column': False,
            'has_category_column': False,
            'has_source_location_column': False,
            'sample_rows': []
        }


# For testing/development
def extract_keywords_from_excel(excel_file: Union[str, io.BytesIO]) -> Dict[str, Any]:
    """
    Extract keywords from Excel file with comprehensive processing for API routes.
    
    Args:
        excel_file: Path to Excel file or BytesIO object from upload
        
    Returns:
        Dictionary with:
        - keywords: List of processed keyword objects
        - source_locations: Mapping of keywords to region configs
        - stats: Processing statistics
        - validation: Validation results
    """
    try:
        # Load raw data
        if isinstance(excel_file, (str, os.PathLike)):
            raw_keywords = load_keywords(excel_file)
        else:
            # Handle BytesIO from file upload
            raw_keywords = _load_from_bytesio(excel_file)
            
        if not raw_keywords:
            raise ValueError("No valid keywords found in Excel file")
        
        # Process keywords and build region configs
        processed_keywords = []
        source_locations = {}
        region_stats = {"global": 0, "include": 0, "exclude": 0}
        
        for item in raw_keywords:
            keyword = item.get('keyword', '').strip()
            sector = item.get('category', '').strip() or item.get('sector', '').strip()
            source_location = item.get('source_location', '').strip()
            
            if not keyword:
                continue
                
            # Parse source location according to spec
            region_config = _parse_source_location(source_location)
            
            # Build processed keyword object
            processed_keyword = {
                'keyword': keyword,
                'sector': sector or None,
                'region_mode': region_config['region_mode'],
                'country': region_config['country']
            }
            
            processed_keywords.append(processed_keyword)
            source_locations[keyword] = region_config
            
            # Update stats
            region_stats[region_config['region_mode'].lower()] += 1
            
            # Log per specification
            logging.info(f"Excel row: {processed_keyword}")
        
        return {
            'keywords': processed_keywords,
            'source_locations': source_locations,
            'stats': {
                'total_keywords': len(processed_keywords),
                'region_breakdown': region_stats,
                'has_sector_column': any(kw.get('sector') for kw in processed_keywords),
                'unique_sectors': list(set(kw.get('sector') for kw in processed_keywords if kw.get('sector')))
            },
            'validation': {
                'valid': True,
                'row_count': len(processed_keywords),
                'columns_detected': ['Keyword', 'Sector', 'Source location'],
                'sample_keywords': processed_keywords[:3]
            }
        }
        
    except Exception as e:
        logging.error(f"Excel extraction failed: {str(e)}")
        return {
            'keywords': [],
            'source_locations': {},
            'stats': {'total_keywords': 0, 'region_breakdown': {}},
            'validation': {'valid': False, 'error': str(e)}
        }


def _load_from_bytesio(excel_file: io.BytesIO) -> List[Dict[str, Optional[str]]]:
    """Load Excel from BytesIO object (for file uploads)."""
    try:
        import pandas as pd
        
        # Reset position to beginning
        excel_file.seek(0)
        
        # Try to read as Excel
        df = pd.read_excel(excel_file, engine='openpyxl')
        return _process_dataframe(df)
        
    except Exception as e:
        logging.error(f"Failed to load BytesIO Excel: {str(e)}")
        raise ValueError(f"Invalid Excel file: {str(e)}")


def _parse_source_location(source_location: str) -> Dict[str, Optional[str]]:
    """
    Parse source location string according to specification.
    
    Rules:
    - blank / NA / null -> region_mode="GLOBAL"
    - "X" -> region_mode="INCLUDE", country="X"
    - "!X" -> region_mode="EXCLUDE", country="X"
    """
    if not source_location or source_location.lower() in ['na', 'null', 'none', '']:
        return {'region_mode': 'GLOBAL', 'country': None}
    
    source_location = source_location.strip()
    
    if source_location.startswith('!'):
        # Exclude mode: "!South Africa" -> EXCLUDE South Africa
        country = source_location[1:].strip()
        return {'region_mode': 'EXCLUDE', 'country': country if country else None}
    else:
        # Include mode: "South Africa" -> INCLUDE South Africa
        return {'region_mode': 'INCLUDE', 'country': source_location}


def create_sample_excel(file_path: str = "sample_keywords.xlsx"):
    """Create a sample Excel file for testing."""
    try:
        import pandas as pd
        
        sample_data = [
            {"Keyword": "Allianz", "Sector": "Short-term Insurance", "Source location": "South Africa"},
            {"Keyword": "Short-term Insurance", "Sector": "Short-term Insurance", "Source location": "South Africa"},
            {"Keyword": "AIG", "Sector": "Short-term Insurance", "Source location": ""},
            {"Keyword": "1st for Women", "Sector": "Short-term Insurance", "Source location": "South Africa"},
            {"Keyword": "Prudential Authority", "Sector": "Short-term Insurance", "Source location": "!South Africa"},
            {"Keyword": "4 Sure Insurance", "Sector": "Short-term Insurance", "Source location": "South Africa"}
        ]
        
        df = pd.DataFrame(sample_data)
        df.to_excel(file_path, index=False)
        return file_path
        
    except ImportError:
        raise ImportError("pandas required to create sample Excel file")


if __name__ == "__main__":
    # Demo usage
    sample_file = create_sample_excel()
    keywords = load_keywords(sample_file)
    
    print(f"Loaded {len(keywords)} keywords:")
    for kw in keywords:
        print(f"  {kw['keyword']} ({kw['category']}) - {kw['source_location'] or 'Global'}")